import os
import django
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'student_management.settings')
django.setup()

def create_student_import_template():
    """Tạo file Excel mẫu để import sinh viên"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách sinh viên"
    
    # Định nghĩa headers và ví dụ dữ liệu
    headers = [
        ('MSSV', 'DH22TIN001'),
        ('Họ', 'Nguyễn Văn'),
        ('Tên', 'An'),
        ('Email', 'nguyenvanan@student.edu.vn'),
        ('Số điện thoại', '0123456789'),
        ('Giới tính', 'male'),  # male/female/other
        ('Ngày sinh', '2002-05-15'),  # YYYY-MM-DD
        ('Địa chỉ', 'Số 1 Đường ABC, Quận 1, TP.HCM'),
        ('Trạng thái', 'active')  # active/inactive
    ]
    
    # Style cho header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Style cho border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Thêm tiêu đề
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "TEMPLATE IMPORT DANH SÁCH SINH VIÊN"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Thêm hướng dẫn
    instructions = [
        "HƯỚNG DẪN SỬ DỤNG:",
        "1. Điền thông tin sinh viên từ dòng 5 trở đi",
        "2. MSSV: Mã sinh viên (bắt buộc, không trùng)",
        "3. Họ và Tên: Bắt buộc",
        "4. Email: Bắt buộc, định dạng email hợp lệ",
        "5. Số điện thoại: Tùy chọn, 10 số",
        "6. Giới tính: male/female/other",
        "7. Ngày sinh: Định dạng YYYY-MM-DD (VD: 2002-05-15)",
        "8. Địa chỉ: Tùy chọn",
        "9. Trạng thái: active/inactive (mặc định: active)",
        "",
        "LƯU Ý: Không xóa hoặc thay đổi dòng header (dòng 4)"
    ]
    
    # Thêm hướng dẫn vào file
    ws.merge_cells('K1:Q12')
    instruction_cell = ws['K1']
    instruction_cell.value = "\n".join(instructions)
    instruction_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    instruction_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    instruction_cell.border = thin_border
    
    # Thêm headers (dòng 4)
    for col, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Thêm dữ liệu mẫu (dòng 5-7)
    sample_data = [
        ['DH22TIN001', 'Nguyễn Văn', 'An', 'nguyenvanan@student.edu.vn', '0123456789', 'male', '2002-05-15', 'Số 1 Đường ABC, Quận 1, TP.HCM', 'active'],
        ['DH22TIN002', 'Trần Thị', 'Bình', 'tranthibinh@student.edu.vn', '0987654321', 'female', '2002-08-20', 'Số 2 Đường XYZ, Quận 2, TP.HCM', 'active'],
        ['DH22TIN003', 'Lê Hoàng', 'Cường', 'lehoangcuong@student.edu.vn', '0369852147', 'male', '2002-03-10', 'Số 3 Đường DEF, Quận 3, TP.HCM', 'active']
    ]
    
    for row_idx, row_data in enumerate(sample_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Tô màu nhạt cho dữ liệu mẫu
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Thêm ghi chú
    ws.cell(row=9, column=1).value = "* Xóa các dòng mẫu (dòng 5-7) và thêm dữ liệu thực của bạn"
    ws.cell(row=9, column=1).font = Font(italic=True, color="FF0000")
    
    # Điều chỉnh độ rộng cột
    column_widths = [15, 15, 10, 30, 15, 10, 12, 40, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Lưu file
    filename = f"student_import_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Đã tạo file template: {filename}")
    return filename


def create_class_management_template():
    """Tạo file Excel để quản lý lớp học"""
    wb = Workbook()
    
    # Sheet 1: Thông tin lớp
    ws1 = wb.active
    ws1.title = "Thông tin lớp"
    
    class_headers = ['Mã lớp', 'Tên lớp', 'Mô tả', 'Giảng viên Email', 'Số lượng tối đa', 'Trạng thái']
    for col, header in enumerate(class_headers, 1):
        cell = ws1.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Dữ liệu mẫu
    ws1.cell(row=2, column=1).value = "IT001"
    ws1.cell(row=2, column=2).value = "Lập trình Web"
    ws1.cell(row=2, column=3).value = "Môn học về phát triển ứng dụng web"
    ws1.cell(row=2, column=4).value = "teacher@nctu.edu.vn"
    ws1.cell(row=2, column=5).value = 50
    ws1.cell(row=2, column=6).value = "active"
    
    # Sheet 2: Phân công sinh viên vào lớp
    ws2 = wb.create_sheet("Phân công sinh viên")
    
    assign_headers = ['Mã lớp', 'MSSV', 'Ngày tham gia', 'Trạng thái']
    for col, header in enumerate(assign_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Dữ liệu mẫu
    sample_assigns = [
        ['IT001', 'DH22TIN001', datetime.now().strftime('%Y-%m-%d'), 'active'],
        ['IT001', 'DH22TIN002', datetime.now().strftime('%Y-%m-%d'), 'active'],
        ['IT001', 'DH22TIN003', datetime.now().strftime('%Y-%m-%d'), 'active']
    ]
    
    for row_idx, row_data in enumerate(sample_assigns, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx).value = value
    
    # Điều chỉnh độ rộng cột
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 5, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    filename = f"class_management_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Đã tạo file template quản lý lớp: {filename}")
    return filename


def create_grade_import_template():
    """Tạo file Excel mẫu để import điểm"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"
    
    # Headers
    headers = ['MSSV', 'Mã lớp', 'Môn học', 'Điểm', 'Loại kiểm tra', 'Học kỳ', 'Năm học']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Dữ liệu mẫu
    sample_grades = [
        ['DH22TIN001', 'IT001', 'Lập trình Web', '8.5', 'midterm', '1', '2024-2025'],
        ['DH22TIN001', 'IT001', 'Lập trình Web', '9.0', 'final', '1', '2024-2025'],
        ['DH22TIN002', 'IT001', 'Lập trình Web', '7.5', 'midterm', '1', '2024-2025'],
        ['DH22TIN002', 'IT001', 'Lập trình Web', '8.0', 'final', '1', '2024-2025']
    ]
    
    for row_idx, row_data in enumerate(sample_grades, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Thêm ghi chú
    ws.cell(row=7, column=1).value = "Ghi chú:"
    ws.cell(row=8, column=1).value = "- Loại kiểm tra: midterm (giữa kỳ), final (cuối kỳ), assignment (bài tập), quiz"
    ws.cell(row=9, column=1).value = "- Điểm: Thang điểm 10"
    ws.cell(row=10, column=1).value = "- Học kỳ: 1, 2, 3 (hè)"
    
    # Điều chỉnh độ rộng cột
    column_widths = [15, 10, 25, 10, 15, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    filename = f"grade_import_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Đã tạo file template nhập điểm: {filename}")
    return filename


def main():
    """Tạo tất cả các template"""
    print("=" * 60)
    print("TẠO CÁC FILE EXCEL TEMPLATE")
    print("=" * 60)
    
    print("\n1. Tạo template import sinh viên...")
    student_file = create_student_import_template()
    
    print("\n2. Tạo template quản lý lớp...")
    class_file = create_class_management_template()
    
    print("\n3. Tạo template nhập điểm...")
    grade_file = create_grade_import_template()
    
    print("\n" + "=" * 60)
    print("HOÀN TẤT!")
    print("=" * 60)
    print("\n📁 Các file đã tạo:")
    print(f"  1. {student_file} - Template import sinh viên")
    print(f"  2. {class_file} - Template quản lý lớp")
    print(f"  3. {grade_file} - Template nhập điểm")
    print("\n📌 Hướng dẫn sử dụng:")
    print("  1. Mở file Excel template")
    print("  2. Xóa dữ liệu mẫu và điền dữ liệu thực của bạn")
    print("  3. Lưu file và upload lên hệ thống qua chức năng Import")
    print("\n⚠️ Lưu ý:")
    print("  - Giữ nguyên dòng header (dòng 1)")
    print("  - Điền đúng định dạng dữ liệu theo hướng dẫn")
    print("  - MSSV không được trùng lặp")


if __name__ == "__main__":
    main()