#!/usr/bin/env python
"""
Create a simple test Excel file that matches the user's format
"""
import openpyxl
from datetime import datetime

def create_test_excel_file():
    """Create a test Excel file with student data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Title row 1
    ws['A1'] = 'Thông tin sinh viên'
    
    # Headers in row 2 (matching the user's table)
    ws['A2'] = 'STT'
    ws['B2'] = 'Mã sinh viên'
    ws['C2'] = 'Họ đệm'  
    ws['D2'] = 'Tên'
    ws['E2'] = 'Giới tính'
    ws['F2'] = 'Ngày sinh'
    ws['G2'] = 'Lớp học'
    
    # Test data using new unique IDs
    test_data = [
        ('1', 'TEST001', 'Lê Văn', 'Nhựt', 'Nam', '30/10/2004', 'DH22TIN06'),
        ('2', 'TEST002', 'Trần Nguyễn', 'Phương Anh', 'Nữ', '07/11/2004', 'DH22TIN06'),
        ('3', 'TEST003', 'Nguyễn', 'Xuân Bách', 'Nam', '20/10/2004', 'DH22TIN06'),
    ]
    
    # Add data starting from row 3
    for i, data in enumerate(test_data, start=3):
        for j, value in enumerate(data, start=1):
            ws.cell(row=i, column=j).value = value
    
    # Save file
    filename = 'test_students.xlsx'
    wb.save(filename)
    print(f"Created test Excel file: {filename}")
    return filename

if __name__ == "__main__":
    create_test_excel_file()