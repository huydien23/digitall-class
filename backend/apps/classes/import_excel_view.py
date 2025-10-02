from rest_framework.decorators import api_view, permission_classes
from rest_framework import permissions, status
from rest_framework.response import Response
from django.utils import timezone

from .models import Class, ClassStudent
from apps.students.models import Student
from django.conf import settings
from apps.students.utils import build_student_email

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_students_from_excel(request, class_id):
    """Import students from Excel file to a specific class"""
    try:
        # Get the class
        class_obj = Class.objects.get(id=class_id)

        # Check permission - must be the teacher of this class
        if class_obj.teacher != request.user:
            return Response(
                {'error': 'You can only import students to your own classes'},
                status=status.HTTP_403_FORBIDDEN
            )

        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']

        # Validate file extension and basic MIME/signature
        try:
            from apps.core.validators import validate_document_upload
            validate_document_upload(excel_file, allowed_extensions={'xls', 'xlsx'})
        except Exception:
            # Fallback to simple name-based check if validator raises (e.g., missing Pillow)
            if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
                return Response({
                    'success': False,
                    'message': 'Only Excel files (.xlsx, .xls) are supported'
                }, status=status.HTTP_400_BAD_REQUEST)

        # Parse Excel file
        from openpyxl import load_workbook
        import io
        import unicodedata
        from datetime import datetime, date, timedelta

        def _normalize_header(value):
            if value is None:
                return ''
            # Convert to string, strip, lower, replace non-breaking spaces
            s = str(value).replace('\u00A0', ' ').strip().lower()
            # Remove Vietnamese diacritics
            s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
            # Replace spaces with underscores
            s = ' '.join(s.split())  # collapse multiple spaces
            return s.replace(' ', '_')

        def _parse_dob(v):
            if v in (None, ''):
                return date(2000, 1, 1)
            # If already a date/datetime
            if isinstance(v, datetime):
                return v.date()
            try:
                from datetime import date as _date
                if isinstance(v, _date):
                    return v
            except Exception:
                pass
            # Excel serial number
            try:
                serial = float(v)
                excel_base = datetime(1899, 12, 30)  # Excel 1900 system (accounts for 1900 leap bug)
                return (excel_base + timedelta(days=serial)).date()
            except Exception:
                pass
            # String parsing
            s = str(v).strip()
            # dd/mm/yyyy or dd-mm-yyyy
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    continue
            # Fallback
            return date(2000, 1, 1)

        workbook = load_workbook(io.BytesIO(excel_file.read()))
        worksheet = workbook.active

        # Find header row (scan first 15 rows)
        header_row = None
        for i in range(1, min(16, worksheet.max_row + 1)):
            row = [cell.value for cell in worksheet[i]]
            # Check if this row contains typical headers
            row_str = ' '.join(str(v).lower() for v in row if v)
            if any(term in row_str for term in ['mssv', 'student', 'sinh viên', 'họ', 'tên', 'name']):
                header_row = i
                break

        if not header_row:
            header_row = 2  # Default to row 2 if no header found

        # Get headers (normalized)
        headers = []
        for cell in worksheet[header_row]:
            headers.append(_normalize_header(cell.value))

        # Column mapping for Vietnamese Excel
        column_mapping = {
            'student_id': ['student_id', 'mssv', 'id', 'mã_sinh_viên', 'ma_sinh_vien', 'masv'],
            'full_name': ['họ_tên', 'họ_và_tên', 'ho_ten', 'ho_va_ten', 'fullname', 'full_name', 'ho ten', 'ho va ten', 'full name'],
            'last_name': ['họ_đệm', 'họ', 'ho_dem', 'ho', 'last_name', 'surname', 'ho dem'],
            'first_name': ['tên', 'ten', 'first_name', 'name', 'firstname'],
            'gender': ['giới_tính', 'gioi_tinh', 'gender', 'sex', 'gt'],
            'date_of_birth': ['ngày_sinh', 'ngay_sinh', 'date_of_birth', 'birthday', 'dob', 'ns', 'ngay sinh'],
            'email': ['email', 'mail'],
            'phone': ['phone', 'sdt', 'số_điện_thoại', 'so_dien_thoai', 'điện_thoại', 'dien_thoai', 'so dien thoai'],
        }

        # Map headers to fields (normalize both sides)
        field_mapping = {}
        for field, possible_names in column_mapping.items():
            normalized_possible = {_normalize_header(n) for n in possible_names}
            for idx, header in enumerate(headers):
                if header and header in normalized_possible:
                    field_mapping[field] = idx
                    break

        created_students = []
        added_to_class = []
        errors = []

        # Process data rows
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            try:
                # Extract row data
                row_data = {}
                for field, col_index in field_mapping.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    if field == 'student_id':
                        row_data[field] = (str(cell_value).strip().upper() if cell_value is not None else '')
                    elif field == 'date_of_birth':
                        row_data[field] = _parse_dob(cell_value)
                    else:
                        row_data[field] = (str(cell_value).strip() if cell_value is not None else '')

                # Skip empty rows
                if not row_data.get('student_id'):
                    continue

                # Handle name splitting
                if row_data.get('full_name') and not row_data.get('first_name'):
                    # Split full name into first and last name
                    parts = row_data['full_name'].split()
                    if len(parts) >= 2:
                        row_data['first_name'] = parts[-1]  # Last word is first name
                        row_data['last_name'] = ' '.join(parts[:-1])  # Rest is last name
                    else:
                        row_data['first_name'] = row_data['full_name']
                        row_data['last_name'] = ''
                elif row_data.get('last_name') and row_data.get('first_name'):
                    # Handle case where "Tên" column contains multiple words
                    first_parts = row_data.get('first_name', '').split()
                    if len(first_parts) > 1:
                        # First word is actual first name, rest might be middle name
                        actual_first = first_parts[0]
                        additional_middle = ' '.join(first_parts[1:])
                        row_data['first_name'] = actual_first
                        if additional_middle and row_data.get('last_name'):
                            row_data['last_name'] = row_data['last_name'] + ' ' + additional_middle

                # Convert gender
                if row_data.get('gender'):
                    gender = row_data['gender'].lower()
                    gender_map = {
                        'nam': 'male', 'nữ': 'female', 'nu': 'female',
                        'male': 'male', 'female': 'female',
                        'm': 'male', 'f': 'female', '1': 'male', '0': 'female'
                    }
                    row_data['gender'] = gender_map.get(gender, 'male')
                else:
                    row_data['gender'] = 'male'

                # Set defaults
                if not row_data.get('email'):
                    row_data['email'] = build_student_email(
                        given_name=row_data.get('first_name', ''),
                        student_id=row_data['student_id'],
                        domain=settings.STUDENT_EMAIL_DOMAIN,
                    )
                if not row_data.get('phone'):
                    row_data['phone'] = ''
                if not isinstance(row_data.get('date_of_birth'), date):
                    row_data['date_of_birth'] = _parse_dob(row_data.get('date_of_birth'))

                # Create or update student
                student, created = Student.objects.update_or_create(
                    student_id=row_data['student_id'],
                    defaults={
                        'first_name': row_data.get('first_name', ''),
                        'last_name': row_data.get('last_name', ''),
                        'email': row_data.get('email', ''),
                        'phone': row_data.get('phone', ''),
                        'gender': row_data.get('gender', 'male'),
                        'date_of_birth': row_data.get('date_of_birth') or date(2000, 1, 1),
                        'is_active': True
                    }
                )

                if created:
                    created_students.append(student.student_id)

                # Add student to class
                class_student, cs_created = ClassStudent.objects.get_or_create(
                    class_obj=class_obj,
                    student=student,
                    defaults={
                        'status': ClassStudent.Status.ACTIVE,
                        'is_active': True,
                        'joined_at': timezone.now(),
                        'source': ClassStudent.Source.IMPORT
                    }
                )

                if cs_created:
                    added_to_class.append(student.student_id)

            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e),
                    'data': row_data.get('student_id', f'Row {row_num}')
                })

        return Response({
            'success': len(errors) == 0,
            'message': f'Imported {len(added_to_class)} students to class {class_obj.class_name}',
            'created_students': len(created_students),
            'added_to_class': len(added_to_class),
            'errors': errors,
            'details': {
                'new_students': created_students,
                'added_students': added_to_class
            }
        })

    except Class.DoesNotExist:
        return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        import traceback
        print(f"ERROR: Import failed: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
