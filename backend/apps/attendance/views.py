from rest_framework import generics, status, permissions
import logging
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
import uuid
import qrcode
import io
import base64
from datetime import datetime, timedelta
from .models import Attendance, AttendanceSession
from .serializers import AttendanceSerializer, AttendanceSessionSerializer, AttendanceSessionCreateSerializer, AttendanceCreateSerializer


logger = logging.getLogger(__name__)

class AttendanceListCreateView(generics.ListCreateAPIView):
    """List and create attendance records"""
    queryset = Attendance.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        # Use write serializer for POST to support manual attendance creation
        if self.request.method == 'POST':
            return AttendanceCreateSerializer
        return AttendanceSerializer
    
    def get_queryset(self):
        queryset = Attendance.objects.select_related(
            'session', 'student'
        )
        session_id = self.request.query_params.get('session_id')
        student_id = self.request.query_params.get('student_id')
        
        # Filter by session if provided
        if session_id:
            queryset = queryset.filter(session_id=session_id)
        
        # Filter by student if provided
        if student_id:
            sid = str(student_id).strip()
            logger.debug(f"Filtering attendance by student_id: {sid} (type: {type(student_id)})")
            
            # Always try student_code first (more common case for frontend)
            try:
                queryset = queryset.filter(student__student_id=sid)
                logger.debug(f"Filtered by student__student_id: {sid}, found {queryset.count()} records")
            except Exception as e:
                logger.warning(f"Failed to filter by student_code {sid}: {e}")
                # Fallback: if it's pure numeric, try as Student.pk
                if sid.isdigit():
                    try:
                        queryset = queryset.filter(student_id=int(sid))
                        logger.debug(f"Fallback: filtered by student_id (pk): {sid}")
                    except Exception as e2:
                        logger.error(f"Failed to filter by both student_code and pk for {sid}: {e2}")
            
        return queryset.order_by('-created_at')

    def post(self, request, *args, **kwargs):
        """Robust POST handler that normalizes payload keys and always uses the write serializer.
        This avoids 500 errors if a client accidentally posts without the exact field names.
        """
        try:
            data = request.data.copy()
            logger.debug("Manual attendance POST payload (raw=%s)", dict(request.data))
            # Normalize common aliases
            if 'sessionId' in data and 'session_id' not in data:
                data['session_id'] = data.get('sessionId')
            if 'studentId' in data and 'student_id' not in data:
                data['student_id'] = data.get('studentId')
            if 'student_code' in data and 'student_id' not in data:
                data['student_id'] = data.get('student_code')
            logger.debug("Normalized payload=%s", dict(data))

            if not data.get('session_id'):
                return Response({'error': 'Thiếu session_id trong payload'}, status=status.HTTP_400_BAD_REQUEST)
            if not (data.get('student_id') or data.get('student') or data.get('student_pk')):
                return Response({'error': 'Thiếu student_id hoặc student (pk)'}, status=status.HTTP_400_BAD_REQUEST)

            serializer = AttendanceCreateSerializer(data=data)
            if not serializer.is_valid():
                logger.info("Manual attendance validation errors: %s", serializer.errors)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            attendance = serializer.save()
            return Response(AttendanceSerializer(attendance).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error("Error in attendance POST: %s", str(e), exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request, *args, **kwargs):
        # Create attendance with write serializer but return full read serializer
        write_serializer = self.get_serializer(data=request.data)
        write_serializer.is_valid(raise_exception=True)
        attendance = write_serializer.save()
        read_serializer = AttendanceSerializer(attendance)
        headers = self.get_success_headers(write_serializer.data)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class AttendanceDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete an attendance record"""
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]


class AttendanceSessionListCreateView(generics.ListCreateAPIView):
    """List and create attendance sessions"""
    queryset = AttendanceSession.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return AttendanceSessionCreateSerializer
        return AttendanceSessionSerializer

    def get_queryset(self):
        queryset = AttendanceSession.objects.select_related(
            'class_obj', 'created_by'
        ).prefetch_related('attendances')
        class_id = self.request.query_params.get('class_id', None)
        session_type = self.request.query_params.get('session_type', None)
        group_name = self.request.query_params.get('group_name', None)
        session_date = self.request.query_params.get('session_date', None)
        
        if class_id is not None:
            queryset = queryset.filter(class_obj_id=class_id)
        if session_type is not None:
            queryset = queryset.filter(session_type=session_type)
        if group_name is not None:
            queryset = queryset.filter(group_name=group_name)
        if session_date is not None:
            queryset = queryset.filter(session_date=session_date)
        
        return queryset.order_by('-session_date', '-start_time')


class AttendanceSessionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete an attendance session"""
    queryset = AttendanceSession.objects.all()
    serializer_class = AttendanceSessionSerializer
    permission_classes = [permissions.IsAuthenticated]

    # FE hiện đang gọi PUT với payload tối giản (vd: {"is_active": false}).
    # Cho phép PUT hoạt động như PATCH (partial update) để tránh 400.
    def put(self, request, *args, **kwargs):
        kwargs.setdefault('partial', True)
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete an attendance session with permission check.
        Only the class teacher or admin can delete it.
        Returns info about deleted related attendances.
        """
        try:
            session = self.get_object()
            # Permission check
            if request.user.role != 'admin' and session.class_obj.teacher != request.user:
                return Response(
                    {'error': 'Bạn không có quyền xóa buổi điểm danh này'},
                    status=status.HTTP_403_FORBIDDEN
                )
            # Count related attendances before deletion
            deleted_attendances = session.attendances.count()
            self.perform_destroy(session)
            return Response({
                'message': 'Đã xóa buổi học',
                'deleted_attendances': deleted_attendances
            }, status=status.HTTP_200_OK)
        except AttendanceSession.DoesNotExist:
            return Response({'error': 'Không tìm thấy buổi điểm danh'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def end_attendance_session(request, session_id):
    """End an attendance session (set is_active=False and set end_time to now if missing)."""
    try:
        session = AttendanceSession.objects.get(id=session_id)
        if request.user.role != 'admin' and session.class_obj.teacher != request.user:
            return Response(
                {'error': 'Bạn không có quyền kết thúc buổi điểm danh này'},
                status=status.HTTP_403_FORBIDDEN
            )
        now = timezone.localtime()
        if not session.end_time:
            session.end_time = now.time()
        session.is_active = False
        session.save(update_fields=['is_active', 'end_time', 'updated_at'])
        return Response({
            'message': 'Đã kết thúc buổi điểm danh',
            'session': AttendanceSessionSerializer(session).data
        })
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'Không tìm thấy buổi điểm danh'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def attendance_statistics(request):
    """Get attendance statistics"""
    total_sessions = AttendanceSession.objects.count()
    total_attendances = Attendance.objects.count()
    present_count = Attendance.objects.filter(status='present').count()
    
    return Response({
        'total_sessions': total_sessions,
        'total_attendances': total_attendances,
        'present_count': present_count,
        'attendance_rate': round((present_count / total_attendances * 100), 2) if total_attendances > 0 else 0,
    })


@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_qr_code(request, session_id):
    """Generate QR code for attendance session"""
    try:
        session = AttendanceSession.objects.get(id=session_id)
        
        # Check permission
        if request.user.role != 'admin' and session.class_obj.teacher != request.user:
            return Response(
                {'error': 'Bạn không có quyền tạo QR code cho buổi điểm danh này'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Generate unique QR code
        qr_code = str(uuid.uuid4())
        session.qr_code = qr_code
        session.save()
        
        # Create QR code image
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_code)
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        expires_dt = timezone.make_aware(datetime.combine(session.session_date, session.end_time)) if session.end_time else None
        return Response({
            'qr_code': qr_code,
            'qr_image': f'data:image/png;base64,{img_str}',
            'session_info': {
                'id': session.id,
                'session_name': session.session_name,
                'class_name': session.class_obj.class_name,
                'session_date': session.session_date,
                'start_time': session.start_time,
                'end_time': session.end_time
            },
            'expires_at': expires_dt
        })
        
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'Không tìm thấy buổi điểm danh'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def check_in_with_qr(request):
    """Check in using QR code"""
    try:
        qr_code = request.data.get('qr_code')
        student_id = request.data.get('student_id')
        
        if not qr_code or not student_id:
            return Response(
                {'error': 'QR code và mã sinh viên là bắt buộc'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Find session by QR code
        session = AttendanceSession.objects.get(qr_code=qr_code, is_active=True)
        
        # Check if session is still active (within time range)
        now = timezone.now()
        session_datetime = timezone.make_aware(
            datetime.combine(session.session_date, session.end_time)
        ) if session.end_time else None
        
        if session_datetime and now > session_datetime:
            return Response(
                {'error': 'Buổi điểm danh đã kết thúc'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Find student
        from apps.students.models import Student
        student = Student.objects.get(student_id=student_id)
        
        # Check if student is in the class
        from apps.classes.models import ClassStudent
        if not ClassStudent.objects.filter(
            class_obj=session.class_obj,
            student=student,
            is_active=True
        ).exists():
            # Auto-enroll if class allows it
            if getattr(session.class_obj, 'allow_auto_enroll_on_attendance_qr', False):
                ClassStudent.objects.get_or_create(
                    class_obj=session.class_obj,
                    student=student,
                    defaults={
                        'status': 'active',
                        'is_active': True,
                        'joined_at': timezone.now(),
                        'source': 'qr'
                    }
                )
            else:
                return Response(
                    {'error': 'Sinh viên không thuộc lớp này'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Check if already attended
        attendance, created = Attendance.objects.get_or_create(
            session=session,
            student=student,
            defaults={
                'status': 'present',
                'check_in_time': now
            }
        )
        
        if not created:
            if attendance.status == 'present':
                return Response(
                    {'error': 'Bạn đã điểm danh rồi'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                # Update existing attendance
                attendance.status = 'present'
                attendance.check_in_time = now
                attendance.save()
        
        return Response({
            'message': 'Điểm danh thành công',
            'attendance': AttendanceSerializer(attendance).data,
            'session_info': {
                'session_name': session.session_name,
                'class_name': session.class_obj.class_name,
                'session_date': session.session_date
            }
        })
        
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'QR code không hợp lệ hoặc buổi điểm danh không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
    except Student.DoesNotExist:
        return Response({'error': 'Không tìm thấy sinh viên'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def attendance_analytics(request, session_id):
    """Get attendance analytics for a session"""
    try:
        session = AttendanceSession.objects.get(id=session_id)
        
        # Check permission
        if request.user.role != 'admin' and session.class_obj.teacher != request.user:
            return Response(
                {'error': 'Bạn không có quyền xem thống kê buổi điểm danh này'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get attendance data
        total_students = session.class_obj.current_students_count
        present_count = Attendance.objects.filter(session=session, status='present').count()
        absent_count = total_students - present_count
        attendance_rate = (present_count / total_students * 100) if total_students > 0 else 0
        
        # Get attendance by time
        attendances = Attendance.objects.filter(session=session, status='present')
        attendance_times = []
        for att in attendances:
            attendance_times.append({
                'student_id': att.student.student_id,
                'student_name': att.student.full_name,
                'check_in_time': att.check_in_time,
                'status': att.status
            })
        
        return Response({
            'session_info': {
                'id': session.id,
                'session_name': session.session_name,
                'class_name': session.class_obj.class_name,
                'session_date': session.session_date,
                'start_time': session.start_time,
                'end_time': session.end_time
            },
            'statistics': {
                'total_students': total_students,
                'present_count': present_count,
                'absent_count': absent_count,
                'attendance_rate': round(attendance_rate, 2)
            },
            'attendance_details': attendance_times
        })
        
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'Không tìm thấy buổi điểm danh'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_excel(request):
    """Import attendance records from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Only Excel files (.xlsx, .xls) are supported'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse Excel file with openpyxl
        from openpyxl import load_workbook
        import io
        
        # Load workbook from uploaded file
        workbook = load_workbook(io.BytesIO(excel_file.read()))
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value.lower().replace(' ', '_') if cell.value else '')
        
        # Expected columns mapping
        column_mapping = {
            'student_id': ['student_id', 'mssv', 'id'],
            'session_id': ['session_id', 'buoi_diem_danh', 'session'],
            'status': ['status', 'trang_thai', 'attendance_status'],
            'check_in_time': ['check_in_time', 'gio_vao', 'time_in'],
            'check_out_time': ['check_out_time', 'gio_ra', 'time_out'],
            'notes': ['notes', 'ghi_chu', 'note']
        }
        
        # Map headers to expected fields
        field_mapping = {}
        for field, possible_names in column_mapping.items():
            for header in headers:
                if header in possible_names:
                    field_mapping[field] = headers.index(header)
                    break
        
        created_attendance = []
        errors = []
        
        # Process each row
        for row_num in range(2, worksheet.max_row + 1):
            try:
                row_data = {}
                for field, col_index in field_mapping.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    row_data[field] = str(cell_value).strip() if cell_value else ''
                
                # Set defaults for missing fields
                if not row_data.get('status'):
                    row_data['status'] = 'present'
                if not row_data.get('check_in_time'):
                    row_data['check_in_time'] = timezone.now().isoformat()
                
                # Validate required fields
                if not row_data.get('student_id') or not row_data.get('session_id'):
                    errors.append({
                        'row': row_num,
                        'error': 'Missing required fields: student_id and session_id',
                        'data': row_data
                    })
                    continue
                
                # Convert session_id to int
                try:
                    row_data['session_id'] = int(row_data['session_id'])
                except ValueError:
                    errors.append({
                        'row': row_num,
                        'error': f'Invalid session_id format: {row_data["session_id"]}',
                        'data': row_data
                    })
                    continue
                
                # Validate status
                valid_statuses = ['present', 'absent', 'late', 'excused']
                if row_data['status'] not in valid_statuses:
                    errors.append({
                        'row': row_num,
                        'error': f'Invalid status: {row_data["status"]}. Must be one of: {valid_statuses}',
                        'data': row_data
                    })
                    continue
                
                # Create attendance record
                serializer = AttendanceCreateSerializer(data=row_data)
                if serializer.is_valid():
                    attendance = serializer.save()
                    created_attendance.append(AttendanceSerializer(attendance).data)
                else:
                    errors.append({
                        'row': row_num,
                        'errors': serializer.errors,
                        'data': row_data
                    })
                    
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e),
                    'data': f'Row {row_num} processing failed'
                })
        
        return Response({
            'success': True,
            'message': f'Successfully imported {len(created_attendance)} attendance records from Excel file',
            'created_count': len(created_attendance),
            'created_attendance': created_attendance,
            'errors': errors,
            'details': {
                'total_rows_processed': worksheet.max_row - 1,
                'successful_imports': len(created_attendance),
                'failed_imports': len(errors)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)