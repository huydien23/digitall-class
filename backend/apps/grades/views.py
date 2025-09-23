from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Q, Avg, Count
from django.http import HttpResponse, JsonResponse
from .models import Grade
from .serializers import GradeSerializer, GradeCreateSerializer


class GradeListCreateView(generics.ListCreateAPIView):
    """List and create grades"""
    queryset = Grade.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return GradeCreateSerializer
        return GradeSerializer
    
    def get_queryset(self):
        queryset = Grade.objects.all()
        student_id = self.request.query_params.get('student_id', None)
        class_id = self.request.query_params.get('class_id', None)
        
        if student_id is not None:
            queryset = queryset.filter(student_id=student_id)
        if class_id is not None:
            queryset = queryset.filter(class_instance_id=class_id)
            
        return queryset.order_by('-created_at')


class GradeDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a grade"""
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [permissions.IsAuthenticated]


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def grade_statistics(request):
    """Get comprehensive grade statistics"""
    try:
        from django.db.models import Avg, Count, Min, Max, Q
        from datetime import date, timedelta
        
        # Basic statistics
        total_grades = Grade.objects.count()
        avg_score = Grade.objects.aggregate(avg_score=Avg('score'))['avg_score']
        min_score = Grade.objects.aggregate(min_score=Min('score'))['min_score']
        max_score = Grade.objects.aggregate(max_score=Max('score'))['max_score']
        
        # Grade distribution by type
        grade_by_type = Grade.objects.values('grade_type').annotate(
            count=Count('id'),
            avg_score=Avg('score')
        ).order_by('-count')
        
        # Grade distribution by letter grade
        letter_grades = ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F']
        grade_distribution = {}
        
        for letter in letter_grades:
            count = Grade.objects.filter(
                score__gte=_get_min_score_for_letter(letter),
                score__lt=_get_max_score_for_letter(letter)
            ).count()
            grade_distribution[letter] = count
        
        # Recent activity
        thirty_days_ago = date.today() - timedelta(days=30)
        recent_grades = Grade.objects.filter(created_at__gte=thirty_days_ago).count()
        
        # Subject performance
        subject_stats = Grade.objects.values('subject__subject_name').annotate(
            count=Count('id'),
            avg_score=Avg('score'),
            min_score=Min('score'),
            max_score=Max('score')
        ).order_by('-avg_score')[:10]
        
        return Response({
            'total_grades': total_grades,
            'average_score': round(avg_score, 2) if avg_score else 0,
            'score_range': {
                'min': float(min_score) if min_score else 0,
                'max': float(max_score) if max_score else 0
            },
            'grade_by_type': list(grade_by_type),
            'grade_distribution': grade_distribution,
            'recent_activity': {
                'grades_last_30_days': recent_grades
            },
            'subject_performance': list(subject_stats)
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _get_min_score_for_letter(letter):
    """Get minimum score for letter grade"""
    grade_ranges = {
        'A+': 9.5, 'A': 8.5, 'A-': 8.0,
        'B+': 7.5, 'B': 7.0, 'B-': 6.5,
        'C+': 6.0, 'C': 5.5, 'C-': 5.0,
        'D+': 4.5, 'D': 4.0, 'D-': 3.5,
        'F': 0
    }
    return grade_ranges.get(letter, 0)


def _get_max_score_for_letter(letter):
    """Get maximum score for letter grade"""
    if letter == 'A+':
        return 10.1
    return _get_min_score_for_letter(letter) + 0.5


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def student_grade_summary(request, student_id):
    """Get comprehensive grade summary for a student"""
    try:
        from apps.students.models import Student
        
        student = Student.objects.get(student_id=student_id)
        grades = Grade.objects.filter(student=student)
        
        # Overall statistics
        total_grades = grades.count()
        avg_score = grades.aggregate(avg=Avg('score'))['avg']
        
        # Grade by subject
        subject_grades = grades.values('subject__subject_name', 'subject__subject_id').annotate(
            count=Count('id'),
            avg_score=Avg('score'),
            latest_grade=Max('created_at')
        ).order_by('-avg_score')
        
        # Grade by type
        type_grades = grades.values('grade_type').annotate(
            count=Count('id'),
            avg_score=Avg('score')
        ).order_by('-avg_score')
        
        # Recent grades
        recent_grades = grades.order_by('-created_at')[:10]
        
        # GPA calculation (simplified)
        gpa = _calculate_gpa(grades)
        
        return Response({
            'student_info': {
                'student_id': student.student_id,
                'full_name': student.full_name,
                'email': student.email
            },
            'overall_stats': {
                'total_grades': total_grades,
                'average_score': round(avg_score, 2) if avg_score else 0,
                'gpa': gpa
            },
            'subject_performance': list(subject_grades),
            'grade_by_type': list(type_grades),
            'recent_grades': [
                {
                    'subject': grade.subject.subject_name,
                    'grade_type': grade.get_grade_type_display(),
                    'score': float(grade.score),
                    'max_score': float(grade.max_score),
                    'percentage': grade.percentage,
                    'letter_grade': grade.letter_grade,
                    'created_at': grade.created_at
                } for grade in recent_grades
            ]
        })
        
    except Student.DoesNotExist:
        return Response({'error': 'Không tìm thấy sinh viên'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def class_grade_summary(request, class_id):
    """Get grade summary for all students in a class"""
    try:
        from apps.classes.models import Class
        
        class_obj = Class.objects.get(id=class_id)
        
        # Check permission
        if request.user.role != 'admin' and class_obj.teacher != request.user:
            return Response(
                {'error': 'Bạn không có quyền xem điểm của lớp này'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all grades for this class
        grades = Grade.objects.filter(class_obj=class_obj)
        
        # Class statistics
        total_grades = grades.count()
        avg_score = grades.aggregate(avg=Avg('score'))['avg']
        
        # Student performance
        student_performance = grades.values(
            'student__student_id', 
            'student__first_name', 
            'student__last_name'
        ).annotate(
            count=Count('id'),
            avg_score=Avg('score'),
            min_score=Min('score'),
            max_score=Max('score')
        ).order_by('-avg_score')
        
        # Subject performance
        subject_performance = grades.values('subject__subject_name').annotate(
            count=Count('id'),
            avg_score=Avg('score'),
            min_score=Min('score'),
            max_score=Max('score')
        ).order_by('-avg_score')
        
        # Grade distribution
        grade_distribution = {}
        for letter in ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F']:
            count = grades.filter(
                score__gte=_get_min_score_for_letter(letter),
                score__lt=_get_max_score_for_letter(letter)
            ).count()
            grade_distribution[letter] = count
        
        return Response({
            'class_info': {
                'class_id': class_obj.class_id,
                'class_name': class_obj.class_name,
                'teacher': f"{class_obj.teacher.first_name} {class_obj.teacher.last_name}",
                'total_students': class_obj.current_students_count
            },
            'class_statistics': {
                'total_grades': total_grades,
                'average_score': round(avg_score, 2) if avg_score else 0,
                'grade_distribution': grade_distribution
            },
            'student_performance': list(student_performance),
            'subject_performance': list(subject_performance)
        })
        
    except Class.DoesNotExist:
        return Response({'error': 'Không tìm thấy lớp học'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _calculate_gpa(grades):
    """Calculate GPA based on grades"""
    if not grades.exists():
        return 0.0
    
    total_points = 0
    total_credits = 0
    
    for grade in grades:
        # Convert score to GPA points
        gpa_points = _score_to_gpa_points(grade.score)
        credits = grade.subject.credits if hasattr(grade.subject, 'credits') else 3
        
        total_points += gpa_points * credits
        total_credits += credits
    
    return round(total_points / total_credits, 2) if total_credits > 0 else 0.0


def _score_to_gpa_points(score):
    """Convert score to GPA points"""
    if score >= 9.0:
        return 4.0
    elif score >= 8.5:
        return 3.7
    elif score >= 8.0:
        return 3.3
    elif score >= 7.5:
        return 3.0
    elif score >= 7.0:
        return 2.7
    elif score >= 6.5:
        return 2.3
    elif score >= 6.0:
        return 2.0
    elif score >= 5.5:
        return 1.7
    elif score >= 5.0:
        return 1.3
    elif score >= 4.5:
        return 1.0
    else:
        return 0.0


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_excel(request):
    """Import grades from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Only Excel files (.xlsx, .xls) are supported'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse Excel file with openpyxl
        from openpyxl import load_workbook
        import io
        
        # Load workbook from uploaded file
        workbook = load_workbook(io.BytesIO(excel_file.read()))
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value.lower().replace(' ', '_') if cell.value else '')
        
        # Expected columns mapping
        column_mapping = {
            'student_id': ['student_id', 'mssv', 'id'],
            'class_id': ['class_id', 'lop', 'class'],
            'subject': ['subject', 'mon_hoc', 'course'],
            'score': ['score', 'diem', 'grade', 'mark'],
            'exam_type': ['exam_type', 'loai_kiem_tra', 'type'],
            'semester': ['semester', 'hoc_ky', 'term'],
            'academic_year': ['academic_year', 'nam_hoc', 'year']
        }
        
        # Map headers to expected fields
        field_mapping = {}
        for field, possible_names in column_mapping.items():
            for header in headers:
                if header in possible_names:
                    field_mapping[field] = headers.index(header)
                    break
        
        created_grades = []
        errors = []
        
        # Process each row
        for row_num in range(2, worksheet.max_row + 1):
            try:
                row_data = {}
                for field, col_index in field_mapping.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    row_data[field] = str(cell_value).strip() if cell_value else ''
                
                # Set defaults for missing fields
                if not row_data.get('exam_type'):
                    row_data['exam_type'] = 'midterm'
                if not row_data.get('semester'):
                    row_data['semester'] = '1'
                if not row_data.get('academic_year'):
                    row_data['academic_year'] = '2024-2025'
                
                # Validate required fields
                if not row_data.get('student_id') or not row_data.get('score'):
                    errors.append({
                        'row': row_num,
                        'error': 'Missing required fields: student_id and score',
                        'data': row_data
                    })
                    continue
                
                # Convert score to float
                try:
                    row_data['score'] = float(row_data['score'])
                except ValueError:
                    errors.append({
                        'row': row_num,
                        'error': f'Invalid score format: {row_data["score"]}',
                        'data': row_data
                    })
                    continue
                
                # Create grade
                serializer = GradeCreateSerializer(data=row_data)
                if serializer.is_valid():
                    grade = serializer.save()
                    created_grades.append(GradeSerializer(grade).data)
                else:
                    errors.append({
                        'row': row_num,
                        'errors': serializer.errors,
                        'data': row_data
                    })
                    
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e),
                    'data': f'Row {row_num} processing failed'
                })
        
        return Response({
            'success': True,
            'message': f'Successfully imported {len(created_grades)} grades from Excel file',
            'created_count': len(created_grades),
            'created_grades': created_grades,
            'errors': errors,
            'details': {
                'total_rows_processed': worksheet.max_row - 1,
                'successful_imports': len(created_grades),
                'failed_imports': len(errors)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)