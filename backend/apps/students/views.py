from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.core.exceptions import ValidationError
from django.db import transaction
import logging

logger = logging.getLogger(__name__)
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from .models import Student
from .serializers import StudentSerializer, StudentCreateSerializer
from .bulk_views import bulk_create_students


class StudentListCreateView(generics.ListCreateAPIView):
    """List and create students with pagination"""
    queryset = Student.objects.select_related('user').all()
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return StudentCreateSerializer
        return StudentSerializer
    
    def get_queryset(self):
        try:
            queryset = Student.objects.select_related('user').all()
            search = self.request.query_params.get('search', None)
            is_active = self.request.query_params.get('is_active', None)
            
            if search is not None:
                queryset = queryset.filter(
                    Q(first_name__icontains=search) |
                    Q(last_name__icontains=search) |
                    Q(student_id__icontains=search) |
                    Q(email__icontains=search)
                )
            
            if is_active is not None:
                is_active = is_active.lower() == 'true'
                queryset = queryset.filter(is_active=is_active)
                
            return queryset.order_by('-created_at')
        except Exception as e:
            logger.error(f'Error in get_queryset: {str(e)}')
            return Student.objects.none()
    
    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            logger.warning(f'Validation error creating student: {str(e)}')
            return Response(
                {'error': 'Validation failed', 'details': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f'Unexpected error creating student: {str(e)}')
            return Response(
                {'error': 'An unexpected error occurred'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a student"""
    queryset = Student.objects.select_related('user').all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'student_id'
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return StudentCreateSerializer
        return StudentSerializer
    
    def retrieve(self, request, *args, **kwargs):
        try:
            return super().retrieve(request, *args, **kwargs)
        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f'Error retrieving student: {str(e)}')
            return Response(
                {'error': 'An unexpected error occurred'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValidationError as e:
            logger.warning(f'Validation error updating student: {str(e)}')
            return Response(
                {'error': 'Validation failed', 'details': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f'Error updating student: {str(e)}')
            return Response(
                {'error': 'An unexpected error occurred'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            # Soft delete instead of hard delete
            instance.is_active = False
            instance.save()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f'Error deleting student: {str(e)}')
            return Response(
                {'error': 'An unexpected error occurred'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_create_students(request):
    """Bulk create students with transaction support"""
    if not isinstance(request.data.get('students'), list):
        return Response(
            {'error': 'Invalid data format. Expected {"students": [...]}'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    students_data = request.data.get('students', [])
    if not students_data:
        return Response(
            {'error': 'No students data provided'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    created_students = []
    errors = []
    
    try:
        with transaction.atomic():
            for i, student_data in enumerate(students_data):
                try:
                    serializer = StudentCreateSerializer(data=student_data)
                    if serializer.is_valid():
                        student = serializer.save()
                        created_students.append(StudentSerializer(student).data)
                        logger.info(f'Successfully created student: {student.student_id}')
                    else:
                        errors.append({
                            'index': i,
                            'data': student_data,
                            'errors': serializer.errors
                        })
                        logger.warning(f'Validation failed for student at index {i}: {serializer.errors}')
                except Exception as e:
                    errors.append({
                        'index': i,
                        'data': student_data,
                        'error': str(e)
                    })
                    logger.error(f'Error creating student at index {i}: {str(e)}')
        
        return Response({
            'success': len(created_students) > 0,
            'created_count': len(created_students),
            'error_count': len(errors),
            'created_students': created_students,
            'errors': errors
        })
    except Exception as e:
        logger.error(f'Bulk create transaction failed: {str(e)}')
        return Response({
            'success': False,
            'error': 'Transaction failed: ' + str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_excel(request):
    """Import students from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Only Excel files (.xlsx, .xls) are supported'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse Excel file with openpyxl
        from openpyxl import load_workbook
        import io
        
        # Load workbook from uploaded file
        workbook = load_workbook(io.BytesIO(excel_file.read()))
        worksheet = workbook.active
        
        # Get headers from second row (row 2) - Vietnamese Excel format
        headers = []
        for cell in worksheet[2]:  # Row 2 instead of row 1
            headers.append(cell.value.lower().replace(' ', '_') if cell.value else '')
        
        # Expected columns mapping for Vietnamese Excel
        column_mapping = {
            'student_id': ['student_id', 'mssv', 'id', 'mã_sinh_viên'],
            'first_name': ['first_name', 'ten', 'ho_ten', 'name', 'tên'],
            'last_name': ['last_name', 'ho', 'surname', 'họ_đệm'],
            'email': ['email', 'mail'],
            'phone': ['phone', 'sdt', 'telephone'],
            'gender': ['gender', 'gioi_tinh', 'sex', 'giới_tính'],
            'date_of_birth': ['date_of_birth', 'ngay_sinh', 'birthday', 'ngày_sinh'],
            'address': ['address', 'dia_chi', 'location']
        }
        
        # Map headers to expected fields
        field_mapping = {}
        for field, possible_names in column_mapping.items():
            for header in headers:
                if header in possible_names:
                    field_mapping[field] = headers.index(header)
                    break
        
        # Debug: Print headers and field mapping
        print(f"DEBUG: Headers found: {headers}")
        print(f"DEBUG: Field mapping: {field_mapping}")
        
        created_students = []
        errors = []
        
        # Process each row starting from row 3 (data starts after header row 2)
        for row_num in range(3, worksheet.max_row + 1):
            try:
                row_data = {}
                for field, col_index in field_mapping.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    # Better handling of empty/null values
                    if cell_value is None or str(cell_value).strip() == '':
                        row_data[field] = ''
                    else:
                        row_data[field] = str(cell_value).strip()
                
                # Debug: Print first few rows
                if row_num <= 5:
                    print(f"DEBUG: Row {row_num} data: {row_data}")
                
                # Set defaults for missing fields
                if not row_data.get('gender'):
                    row_data['gender'] = 'male'
                if not row_data.get('date_of_birth'):
                    row_data['date_of_birth'] = '2000-01-01'
                if not row_data.get('email'):
                    # Generate email from student_id if not provided
                    row_data['email'] = f"{row_data.get('student_id', 'student')}@student.edu.vn"
                if not row_data.get('phone'):
                    row_data['phone'] = ''
                if not row_data.get('address'):
                    row_data['address'] = ''
                
                # Convert Vietnamese gender to English
                if row_data.get('gender'):
                    gender_map = {
                        'nam': 'male',
                        'nữ': 'female',
                        'male': 'male',
                        'female': 'female'
                    }
                    row_data['gender'] = gender_map.get(row_data['gender'].lower(), 'male')
                
                # Convert date format from DD/MM/YYYY to YYYY-MM-DD
                if row_data.get('date_of_birth') and row_data['date_of_birth'] != '2000-01-01':
                    try:
                        from datetime import datetime
                        # Try to parse DD/MM/YYYY format
                        if '/' in row_data['date_of_birth']:
                            date_obj = datetime.strptime(row_data['date_of_birth'], '%d/%m/%Y')
                            row_data['date_of_birth'] = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        # If parsing fails, keep original or set default
                        row_data['date_of_birth'] = '2000-01-01'
                
                # Validate required fields
                if not row_data.get('student_id') or not row_data.get('first_name'):
                    error_msg = f'Missing required fields: student_id={row_data.get("student_id")}, first_name={row_data.get("first_name")}'
                    print(f"DEBUG: Row {row_num} validation failed: {error_msg}")
                    errors.append({
                        'row': row_num,
                        'error': error_msg,
                        'data': row_data
                    })
                    continue
                
                # Check if student already exists
                if Student.objects.filter(student_id=row_data['student_id']).exists():
                    errors.append({
                        'row': row_num,
                        'error': f'Student with ID {row_data["student_id"]} already exists',
                        'data': row_data
                    })
                    continue
                
                # Create student
                serializer = StudentCreateSerializer(data=row_data)
                if serializer.is_valid():
                    student = serializer.save()
                    created_students.append(StudentSerializer(student).data)
                    print(f"DEBUG: Row {row_num} created successfully: {student.student_id}")
                else:
                    print(f"DEBUG: Row {row_num} serializer failed: {serializer.errors}")
                    errors.append({
                        'row': row_num,
                        'error': f'Validation failed: {serializer.errors}',
                        'data': row_data
                    })
                    
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e),
                    'data': f'Row {row_num} processing failed'
                })
        
        # Determine success based on actual results
        success = len(created_students) > 0 or len(errors) == 0
        
        return Response({
            'success': success,
            'message': f'Successfully imported {len(created_students)} students from Excel file' if success else f'Import completed with {len(errors)} errors',
            'created_count': len(created_students),
            'created_students': created_students,
            'errors': errors,
            'details': {
                'total_rows_processed': worksheet.max_row - 2,  # Subtract header row (row 2)
                'successful_imports': len(created_students),
                'failed_imports': len(errors)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_excel(request):
    """Export students to Excel/CSV file"""
    try:
        import csv
        import io
        
        # Get all students
        students = Student.objects.all()
        
        # Create CSV content
        output = io.StringIO()
        fieldnames = [
            'student_id', 'first_name', 'last_name', 'email', 
            'phone', 'gender', 'date_of_birth', 'address', 'is_active'
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for student in students:
            writer.writerow({
                'student_id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'email': student.email,
                'phone': student.phone or '',
                'gender': student.gender,
                'date_of_birth': student.date_of_birth,
                'address': student.address or '',
                'is_active': student.is_active
            })
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv'
        )
        response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
        
        return response
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def student_statistics(request):
    """Get comprehensive student statistics"""
    try:
        from django.db.models import Count, Q
        from datetime import date, timedelta
        
        # Basic counts
        total_students = Student.objects.count()
        active_students = Student.objects.filter(is_active=True).count()
        inactive_students = Student.objects.filter(is_active=False).count()
        
        # Gender distribution
        gender_stats = Student.objects.values('gender').annotate(count=Count('id'))
        
        # Age distribution
        today = date.today()
        age_groups = {
            '18-20': Student.objects.filter(
                date_of_birth__gte=today - timedelta(days=20*365),
                date_of_birth__lt=today - timedelta(days=18*365)
            ).count(),
            '21-25': Student.objects.filter(
                date_of_birth__gte=today - timedelta(days=25*365),
                date_of_birth__lt=today - timedelta(days=21*365)
            ).count(),
            '26-30': Student.objects.filter(
                date_of_birth__gte=today - timedelta(days=30*365),
                date_of_birth__lt=today - timedelta(days=26*365)
            ).count(),
            '30+': Student.objects.filter(
                date_of_birth__lt=today - timedelta(days=30*365)
            ).count(),
        }
        
        # Recent registrations (last 30 days)
        thirty_days_ago = today - timedelta(days=30)
        recent_registrations = Student.objects.filter(
            created_at__gte=thirty_days_ago
        ).count()
        
        # Students with missing information
        missing_phone = Student.objects.filter(phone__isnull=True).count()
        missing_address = Student.objects.filter(address__isnull=True).count()
        
        return Response({
            'total_students': total_students,
            'active_students': active_students,
            'inactive_students': inactive_students,
            'gender_distribution': list(gender_stats),
            'age_groups': age_groups,
            'recent_registrations': recent_registrations,
            'missing_information': {
                'missing_phone': missing_phone,
                'missing_address': missing_address
            },
            'completion_rate': round(
                ((total_students - missing_phone - missing_address) / total_students * 100) 
                if total_students > 0 else 0, 2
            )
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
